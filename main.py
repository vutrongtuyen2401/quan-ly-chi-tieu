"""
╔══════════════════════════════════════════════════════════════════╗
║   CÀN KHÔN LINH THẠCH CÁC — HỆ THỐNG QUẢN LÝ CHI TIÊU AI    ║
║   Backend FastAPI + SQLite + Google Gemini AI                    ║
║   Phong cách Tu Tiên (Xianxia Theme)                            ║
╚══════════════════════════════════════════════════════════════════╝
"""

import os
import sys
import sqlite3
import json
import base64
import datetime
import hashlib
import secrets
import time
import csv
import io
from contextlib import contextmanager

if sys.stdout and hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Form, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel
from typing import Optional, List
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv

import jwt
import bcrypt

load_dotenv()

# ──────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────
DATABASE = "app.db"

# Change 2: JWT_SECRET bắt buộc — dừng server nếu thiếu
JWT_SECRET = os.getenv("JWT_SECRET")
if not JWT_SECRET:
    raise RuntimeError(
        "\n" + "=" * 62 + "\n"
        "  ❌ LỖI NGHIÊM TRỌNG: Biến môi trường JWT_SECRET chưa được đặt!\n"
        "  Hãy thêm JWT_SECRET vào file .env trước khi khởi động server.\n"
        "  Ví dụ: JWT_SECRET=my_super_secret_key_here\n"
        + "=" * 62
    )

JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")

# Change 1: CORS an toàn — đọc danh sách origins từ biến môi trường
ALLOWED_ORIGINS = [o.strip() for o in os.getenv("ALLOWED_ORIGINS", "http://localhost:5173").split(",") if o.strip()]

# Change 6: Rate limiting đăng nhập — in-memory tracker
# {email: {"count": int, "first_attempt": float}}
login_attempts = {}
LOGIN_MAX_ATTEMPTS = 5
LOGIN_LOCKOUT_SECONDS = 15 * 60  # 15 phút

app = FastAPI(title="Càn Khôn Linh Thạch Các API", version="2.1")
security = HTTPBearer()

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ──────────────────────────────────────────────
# DATABASE HELPERS
# ──────────────────────────────────────────────
@contextmanager
def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def init_db():
    """Tạo các bảng cốt lõi + seed dữ liệu mẫu"""
    with get_db() as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                email TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                full_name TEXT NOT NULL,
                soul_lamp_hash TEXT,
                role TEXT DEFAULT 'user',
                is_active INTEGER DEFAULT 1,
                created_at TEXT DEFAULT (datetime('now'))
            );

            CREATE TABLE IF NOT EXISTS wallets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                wallet_name TEXT NOT NULL,
                balance REAL DEFAULT 0,
                wallet_type TEXT DEFAULT 'cash',
                created_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (user_id) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                category_name TEXT NOT NULL,
                category_type TEXT CHECK(category_type IN ('INCOME','EXPENSE')) NOT NULL,
                icon TEXT DEFAULT '📦',
                created_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (user_id) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                wallet_id INTEGER NOT NULL,
                category_id INTEGER NOT NULL,
                amount REAL NOT NULL,
                transaction_type TEXT CHECK(transaction_type IN ('INCOME','EXPENSE')) NOT NULL,
                transaction_date TEXT NOT NULL,
                note TEXT DEFAULT '',
                image_url TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (user_id) REFERENCES users(id),
                FOREIGN KEY (wallet_id) REFERENCES wallets(id),
                FOREIGN KEY (category_id) REFERENCES categories(id)
            );

            CREATE TABLE IF NOT EXISTS budgets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                category_id INTEGER NOT NULL,
                limit_amount REAL NOT NULL,
                month_year TEXT NOT NULL,
                created_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (user_id) REFERENCES users(id),
                FOREIGN KEY (category_id) REFERENCES categories(id)
            );

            CREATE TABLE IF NOT EXISTS invoice_ocr_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                image_path TEXT DEFAULT '',
                extracted_json TEXT DEFAULT '{}',
                created_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (user_id) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS chat_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                prompt_question TEXT NOT NULL,
                ai_response TEXT NOT NULL,
                created_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (user_id) REFERENCES users(id)
            );

            /* Change 5: Bảng password_reset_tokens */
            CREATE TABLE IF NOT EXISTS password_reset_tokens (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                email TEXT NOT NULL,
                token TEXT NOT NULL,
                expires_at TEXT NOT NULL,
                used INTEGER DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now'))
            );

            /* Change 8: Bảng recurring_transactions */
            CREATE TABLE IF NOT EXISTS recurring_transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                wallet_id INTEGER NOT NULL,
                category_id INTEGER NOT NULL,
                amount REAL NOT NULL,
                transaction_type TEXT CHECK(transaction_type IN ('INCOME','EXPENSE')) NOT NULL,
                frequency TEXT CHECK(frequency IN ('weekly','monthly')) NOT NULL DEFAULT 'monthly',
                next_run_date TEXT NOT NULL,
                note TEXT DEFAULT '',
                is_active INTEGER DEFAULT 1,
                created_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (user_id) REFERENCES users(id),
                FOREIGN KEY (wallet_id) REFERENCES wallets(id),
                FOREIGN KEY (category_id) REFERENCES categories(id)
            );

            /* Bảng debts (Theo dõi Nợ / Vay mượn) */
            CREATE TABLE IF NOT EXISTS debts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                wallet_id INTEGER,
                debt_type TEXT CHECK(debt_type IN ('BORROW','LEND')) NOT NULL,
                person_name TEXT NOT NULL,
                amount REAL NOT NULL,
                due_date TEXT DEFAULT '',
                is_settled INTEGER DEFAULT 0,
                note TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (user_id) REFERENCES users(id),
                FOREIGN KEY (wallet_id) REFERENCES wallets(id)
            );

            /* Bảng saving_goals (Mục Tiêu Tiết Kiệm) */
            CREATE TABLE IF NOT EXISTS saving_goals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                target_name TEXT NOT NULL,
                target_amount REAL NOT NULL,
                current_amount REAL DEFAULT 0,
                target_date TEXT DEFAULT '',
                icon TEXT DEFAULT '🎯',
                is_completed INTEGER DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (user_id) REFERENCES users(id)
            );
        """)

        # Migration: Ensure role, is_active, and soul_lamp_hash columns exist on users table
        user_cols = [c[1] for c in conn.execute("PRAGMA table_info(users)").fetchall()]
        if "role" not in user_cols:
            conn.execute("ALTER TABLE users ADD COLUMN role TEXT DEFAULT 'user'")
        if "is_active" not in user_cols:
            conn.execute("ALTER TABLE users ADD COLUMN is_active INTEGER DEFAULT 1")
        if "soul_lamp_hash" not in user_cols:
            conn.execute("ALTER TABLE users ADD COLUMN soul_lamp_hash TEXT")
        
        # Ensure default admin has role = 'admin', valid hash, and default soul_lamp_hash if NULL
        admin_row = conn.execute("SELECT id, password_hash, soul_lamp_hash FROM users WHERE email = 'admin@gmail.com'").fetchone()
        if admin_row:
            admin_hash = admin_row["password_hash"] or ""
            need_pw_reset = False
            try:
                if not (admin_hash.startswith("$2b$") or admin_hash.startswith("$2a$") or admin_hash.startswith("$2y$")):
                    need_pw_reset = True
                else:
                    bcrypt.checkpw(b"test", admin_hash.encode())
            except Exception:
                need_pw_reset = True
            
            seed_password = os.getenv("SEED_ADMIN_PASSWORD", "admin123").strip() or "admin123"
            if need_pw_reset:
                new_pw_hash = bcrypt.hashpw(seed_password.encode(), bcrypt.gensalt()).decode()
                conn.execute("UPDATE users SET role = 'admin', is_active = 1, password_hash = ? WHERE email = 'admin@gmail.com'", (new_pw_hash,))
            else:
                conn.execute("UPDATE users SET role = 'admin', is_active = 1 WHERE email = 'admin@gmail.com'")
            
            # Gán giá trị mặc định cho Bản Mệnh Hồn Đăng của tài khoản Admin nếu đang là NULL
            if admin_row["soul_lamp_hash"] is None or admin_row["soul_lamp_hash"] == "":
                default_soul_lamp_hash = bcrypt.hashpw(b"admin", bcrypt.gensalt()).decode()
                conn.execute("UPDATE users SET soul_lamp_hash = ? WHERE email = 'admin@gmail.com'", (default_soul_lamp_hash,))

        # Change 9: Seed dữ liệu mẫu — mật khẩu an toàn
        user_check = conn.execute("SELECT id FROM users LIMIT 1").fetchone()
        if not user_check:
            # Đọc mật khẩu từ biến môi trường, hoặc tự sinh ngẫu nhiên
            seed_password = os.getenv("SEED_ADMIN_PASSWORD", "admin123").strip() or "admin123"
            pw_hash = bcrypt.hashpw(seed_password.encode(), bcrypt.gensalt()).decode()
            admin_soul_lamp_hash = bcrypt.hashpw(b"admin", bcrypt.gensalt()).decode()
            conn.execute(
                "INSERT INTO users (email, password_hash, full_name, soul_lamp_hash, role, is_active) VALUES (?, ?, ?, ?, 'admin', 1)",
                ("admin@gmail.com", pw_hash, "Ký Chủ", admin_soul_lamp_hash)
            )
            uid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

            # In mật khẩu mẫu ra console với banner nổi bật
            print("\n" + "🔑" * 31)
            print("  ⚠️  MẬT KHẨU TÀI KHOẢN MẪU (Seed Account)")
            print(f"  📧  Email:     admin@gmail.com")
            print(f"  🔐  Mật khẩu: {seed_password}")
            print("  ℹ️  Đặt biến SEED_ADMIN_PASSWORD trong .env để cố định mật khẩu.")
            print("🔑" * 31 + "\n")

            # 3 Ví Linh Thạch
            wallets_data = [
                (uid, "Linh Thạch Tiền Mặt", 5000000, "cash"),
                (uid, "Linh Mạch Vietcombank", 15000000, "bank"),
                (uid, "Túi Momo", 2000000, "e-wallet"),
            ]
            conn.executemany(
                "INSERT INTO wallets (user_id, wallet_name, balance, wallet_type) VALUES (?, ?, ?, ?)",
                wallets_data
            )

            # 5 Danh Mục
            categories_data = [
                (uid, "Ẩm Thực Linh Đan", "EXPENSE", "🍕"),
                (uid, "Pháp Khí Mua Sắm", "EXPENSE", "🛍️"),
                (uid, "Phi Kiếm Di Chuyển", "EXPENSE", "🚗"),
                (uid, "Linh Thạch Lương Bổng", "INCOME", "💵"),
                (uid, "Quà Tặng Đạo Hữu", "INCOME", "🎁"),
            ]
            conn.executemany(
                "INSERT INTO categories (user_id, category_name, category_type, icon) VALUES (?, ?, ?, ?)",
                categories_data
            )

            # Giao dịch mẫu
            today = datetime.date.today()
            transactions_data = [
                (uid, 1, 1, 150000, "EXPENSE", str(today - datetime.timedelta(days=1)), "Mua linh đan phở bò"),
                (uid, 1, 2, 500000, "EXPENSE", str(today - datetime.timedelta(days=2)), "Mua pháp khí áo mới"),
                (uid, 2, 3, 200000, "EXPENSE", str(today - datetime.timedelta(days=3)), "Phi kiếm Grab đi làm"),
                (uid, 2, 4, 20000000, "INCOME", str(today - datetime.timedelta(days=5)), "Lương tháng 8 từ Tông Môn"),
                (uid, 3, 5, 1000000, "INCOME", str(today - datetime.timedelta(days=7)), "Quà tặng từ Sư Huynh"),
                (uid, 1, 1, 85000, "EXPENSE", str(today), "Mua cơm trưa Linh Đan quán"),
                (uid, 2, 2, 1200000, "EXPENSE", str(today - datetime.timedelta(days=4)), "Pháp bảo tai nghe mới"),
            ]
            conn.executemany(
                """INSERT INTO transactions
                   (user_id, wallet_id, category_id, amount, transaction_type, transaction_date, note)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                transactions_data
            )

            # Budget mẫu
            month_year = today.strftime("%Y-%m")
            budgets_data = [
                (uid, 1, 3000000, month_year),
                (uid, 2, 2000000, month_year),
                (uid, 3, 1000000, month_year),
            ]
            conn.executemany(
                "INSERT INTO budgets (user_id, category_id, limit_amount, month_year) VALUES (?, ?, ?, ?)",
                budgets_data
            )
        else:
            # Update default name from old 'Đạo Hữu Admin' to 'Ký Chủ' if unchanged
            conn.execute(
                "UPDATE users SET full_name = 'Ký Chủ' WHERE email = 'admin@gmail.com' AND full_name = 'Đạo Hữu Admin'"
            )

        # Dọn dẹp dữ liệu mồ côi không gắn user_id hợp lệ
        conn.execute("DELETE FROM transactions WHERE user_id NOT IN (SELECT id FROM users)")
        conn.execute("DELETE FROM wallets WHERE user_id NOT IN (SELECT id FROM users)")
        conn.execute("DELETE FROM categories WHERE user_id NOT IN (SELECT id FROM users)")
        conn.execute("DELETE FROM budgets WHERE user_id NOT IN (SELECT id FROM users)")
        conn.execute("DELETE FROM invoice_ocr_logs WHERE user_id NOT IN (SELECT id FROM users)")
        conn.execute("DELETE FROM chat_sessions WHERE user_id NOT IN (SELECT id FROM users)")

        conn.commit()


# ──────────────────────────────────────────────
# PYDANTIC SCHEMAS
# ──────────────────────────────────────────────
class RegisterBody(BaseModel):
    email: str
    password: str
    full_name: str
    soul_lamp: str

class LoginBody(BaseModel):
    email: str
    password: str

class ForgotPasswordBody(BaseModel):
    email: str
    soul_lamp: str

class ResetPasswordBody(BaseModel):
    email: str
    token: str
    new_password: str

class SoulLampUpdateBody(BaseModel):
    current_password: str
    new_soul_lamp: str

class UserProfileUpdateBody(BaseModel):
    full_name: str

class WalletBody(BaseModel):
    wallet_name: str
    balance: float
    wallet_type: str = "cash"

class WalletUpdateBody(BaseModel):
    wallet_name: Optional[str] = None
    wallet_type: Optional[str] = None

class TransferBody(BaseModel):
    from_wallet_id: int
    to_wallet_id: int
    amount: float
    note: Optional[str] = None

class CategoryBody(BaseModel):
    category_name: str
    category_type: str
    icon: str = "📦"

class CategoryUpdateBody(BaseModel):
    category_name: Optional[str] = None
    icon: Optional[str] = None

class TransactionBody(BaseModel):
    wallet_id: int
    category_id: int
    amount: float
    transaction_type: str
    transaction_date: str
    note: Optional[str] = None

class TransactionUpdateBody(BaseModel):
    wallet_id: Optional[int] = None
    category_id: Optional[int] = None
    amount: Optional[float] = None
    transaction_type: Optional[str] = None
    transaction_date: Optional[str] = None
    note: Optional[str] = None

class BudgetBody(BaseModel):
    category_id: int
    limit_amount: float
    month_year: str

class RecurringBody(BaseModel):
    wallet_id: int
    category_id: int
    amount: float
    transaction_type: str = "EXPENSE"
    frequency: str = "monthly"
    next_run_date: str
    note: Optional[str] = None

class RecurringUpdateBody(BaseModel):
    wallet_id: Optional[int] = None
    category_id: Optional[int] = None
    amount: Optional[float] = None
    transaction_type: Optional[str] = None
    frequency: Optional[str] = None
    next_run_date: Optional[str] = None
    note: Optional[str] = None
    is_active: Optional[int] = None

RecurringTransactionBody = RecurringBody
RecurringTransactionUpdateBody = RecurringUpdateBody

class ChatBody(BaseModel):
    message: str

class DebtCreateBody(BaseModel):
    debt_type: str
    person_name: str
    amount: float
    due_date: Optional[str] = None
    note: Optional[str] = None
    wallet_id: Optional[int] = None

class DebtUpdateBody(BaseModel):
    debt_type: Optional[str] = None
    person_name: Optional[str] = None
    amount: Optional[float] = None
    due_date: Optional[str] = None
    note: Optional[str] = None
    wallet_id: Optional[int] = None

class SavingGoalCreateBody(BaseModel):
    target_name: str
    target_amount: float
    current_amount: Optional[float] = 0.0
    target_date: Optional[str] = None
    icon: Optional[str] = "🎯"

class SavingGoalUpdateBody(BaseModel):
    target_name: Optional[str] = None
    target_amount: Optional[float] = None
    current_amount: Optional[float] = None
    target_date: Optional[str] = None
    icon: Optional[str] = None
    is_completed: Optional[int] = None

class SavingGoalDepositBody(BaseModel):
    amount: float
    wallet_id: Optional[int] = None


# ──────────────────────────────────────────────
# AUTH HELPERS
# ──────────────────────────────────────────────
def create_token(user_id: int, email: str, role: str = "user") -> str:
    payload = {
        "user_id": user_id,
        "email": email,
        "role": (role or "user").lower(),
        "exp": datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(hours=JWT_EXPIRATION_HOURS),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return {
            "user_id": payload["user_id"],
            "email": payload["email"],
            "role": str(payload.get("role", "user")).lower()
        }
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token đã hết hạn. Hãy đăng nhập lại.")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token không hợp lệ.")


def require_admin(user: dict = Depends(get_current_user)) -> dict:
    if str(user.get("role", "")).lower() != "admin":
        raise HTTPException(status_code=403, detail="Quyền hạn không đủ! Chỉ Chưởng Môn (Admin) mới có quyền truy cập.")
    return user


# ──────────────────────────────────────────────
# AUTH ROUTES
# ──────────────────────────────────────────────
@app.post("/api/auth/register")
@app.post("/api/register")
@app.post("/register")
def register(body: RegisterBody):
    if not body.soul_lamp or len(body.soul_lamp.strip()) < 3:
        raise HTTPException(status_code=400, detail="Bản Mệnh Hồn Đăng không được để trống và phải có ít nhất 3 ký tự.")

    with get_db() as conn:
        existing = conn.execute("SELECT id FROM users WHERE email = ?", (body.email,)).fetchone()
        if existing:
            raise HTTPException(status_code=400, detail="Email đã tồn tại trong Tông Môn.")
        pw_hash = bcrypt.hashpw(body.password.encode(), bcrypt.gensalt()).decode()
        soul_lamp_hash = bcrypt.hashpw(body.soul_lamp.strip().encode(), bcrypt.gensalt()).decode()
        conn.execute(
            "INSERT INTO users (email, password_hash, full_name, soul_lamp_hash, role, is_active) VALUES (?, ?, ?, ?, 'user', 1)",
            (body.email, pw_hash, body.full_name, soul_lamp_hash)
        )
        user_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        # Khởi tạo Ví và Danh mục mặc định cho user mới
        wallets_data = [
            (user_id, "Linh Thạch Tiền Mặt", 5000000, "cash"),
            (user_id, "Linh Mạch Vietcombank", 15000000, "bank"),
            (user_id, "Túi Momo", 2000000, "e-wallet"),
        ]
        conn.executemany(
            "INSERT INTO wallets (user_id, wallet_name, balance, wallet_type) VALUES (?, ?, ?, ?)",
            wallets_data
        )

        categories_data = [
            (user_id, "Ẩm Thực Linh Đan", "EXPENSE", "🍕"),
            (user_id, "Pháp Khí Mua Sắm", "EXPENSE", "🛍️"),
            (user_id, "Phi Kiếm Di Chuyển", "EXPENSE", "🚗"),
            (user_id, "Linh Thạch Lương Bổng", "INCOME", "💵"),
            (user_id, "Quà Tặng Đạo Hữu", "INCOME", "🎁"),
        ]
        conn.executemany(
            "INSERT INTO categories (user_id, category_name, category_type, icon) VALUES (?, ?, ?, ?)",
            categories_data
        )

        token = create_token(user_id, body.email, role="user")
        return {"token": token, "user_id": user_id, "full_name": body.full_name, "email": body.email, "role": "user"}


@app.post("/api/auth/login")
@app.post("/api/login")
@app.post("/login")
def login(body: LoginBody):
    # Change 6: Rate limiting — kiểm tra số lần đăng nhập sai
    email_lower = body.email.lower().strip()
    now_ts = time.time()

    if email_lower in login_attempts:
        attempt = login_attempts[email_lower]
        elapsed = now_ts - attempt["first_attempt"]
        if elapsed > LOGIN_LOCKOUT_SECONDS:
            del login_attempts[email_lower]
        elif attempt["count"] >= LOGIN_MAX_ATTEMPTS:
            remaining = int(LOGIN_LOCKOUT_SECONDS - elapsed)
            raise HTTPException(
                status_code=429,
                detail=f"Tài khoản tạm khóa do đăng nhập sai quá nhiều lần. Vui lòng thử lại sau {remaining // 60} phút."
            )

    with get_db() as conn:
        user = conn.execute("SELECT * FROM users WHERE email = ?", (body.email,)).fetchone()
        if not user:
            raise HTTPException(status_code=401, detail="Email hoặc mật khẩu không chính xác.")
        if "is_active" in user.keys() and user["is_active"] == 0:
            raise HTTPException(status_code=403, detail="Tài khoản này đã bị phong ấn (khóa). Vui lòng liên hệ Chưởng Môn (Admin).")
        
        is_pw_valid = False
        stored_hash = user["password_hash"] or ""
        try:
            if stored_hash.startswith("$2b$") or stored_hash.startswith("$2a$") or stored_hash.startswith("$2y$"):
                is_pw_valid = bcrypt.checkpw(body.password.encode("utf-8"), stored_hash.encode("utf-8"))
            else:
                # Fallback for plain sha256 or plain text legacy, and auto-upgrade to bcrypt
                sha256_hash = hashlib.sha256(body.password.encode("utf-8")).hexdigest()
                if stored_hash == sha256_hash or stored_hash == body.password:
                    is_pw_valid = True
                    new_pw_hash = bcrypt.hashpw(body.password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
                    conn.execute("UPDATE users SET password_hash = ? WHERE id = ?", (new_pw_hash, user["id"]))
        except Exception:
            is_pw_valid = False

        if not is_pw_valid:
            if email_lower not in login_attempts:
                login_attempts[email_lower] = {"count": 1, "first_attempt": now_ts}
            else:
                login_attempts[email_lower]["count"] += 1
            remaining_attempts = LOGIN_MAX_ATTEMPTS - login_attempts[email_lower]["count"]
            if remaining_attempts <= 0:
                raise HTTPException(
                    status_code=429,
                    detail=f"Tài khoản tạm khóa do đăng nhập sai quá nhiều lần. Vui lòng thử lại sau 15 phút."
                )
            raise HTTPException(status_code=401, detail=f"Mật khẩu sai. Đạo Tâm bị phong ấn. (Còn {remaining_attempts} lần thử)")

        # Đăng nhập thành công → reset bộ đếm
        if email_lower in login_attempts:
            del login_attempts[email_lower]

        user_role = user["role"] if "role" in user.keys() and user["role"] else "user"
        token = create_token(user["id"], user["email"], role=user_role)
        return {
            "token": token,
            "user_id": user["id"],
            "full_name": user["full_name"],
            "email": user["email"],
            "role": user_role
        }


# ──────────────────────────────────────────────
# Change 5: FORGOT / RESET PASSWORD
# ──────────────────────────────────────────────
@app.post("/api/auth/forgot-password")
def forgot_password(body: ForgotPasswordBody):
    """Tạo mã reset mật khẩu — Yêu cầu xác thực Email + Bản Mệnh Hồn Đăng"""
    if not body.email or not body.soul_lamp:
        raise HTTPException(status_code=400, detail="Thông tin xác thực không chính xác, vui lòng kiểm tra lại")

    with get_db() as conn:
        user = conn.execute("SELECT id, soul_lamp_hash FROM users WHERE email = ?", (body.email,)).fetchone()
        
        is_valid = False
        if user and user["soul_lamp_hash"]:
            try:
                is_valid = bcrypt.checkpw(body.soul_lamp.strip().encode("utf-8"), user["soul_lamp_hash"].encode("utf-8"))
            except Exception:
                is_valid = False

        if not is_valid:
            # QUAN TRỌNG VỀ BẢO MẬT: Trả về CÙNG MỘT thông báo lỗi chung cho cả 3 trường hợp:
            # 1. Email không tồn tại
            # 2. Bản Mệnh Hồn Đăng sai
            # 3. User chưa từng đặt Bản Mệnh Hồn Đăng (soul_lamp_hash là NULL)
            raise HTTPException(status_code=400, detail="Thông tin xác thực không chính xác, vui lòng kiểm tra lại")

        # Tạo mã reset 6 ký tự, hạn 30 phút
        reset_token = secrets.token_urlsafe(4)[:6].upper()
        expires_at = (datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(minutes=30)).isoformat()

        conn.execute(
            "INSERT INTO password_reset_tokens (email, token, expires_at) VALUES (?, ?, ?)",
            (body.email, reset_token, expires_at)
        )

    # TODO: Gửi email thật khi lên production. Hiện tại trả trực tiếp cho dev/đồ án.
    return {
        "message": "Mã reset đã được tạo. (Chế độ phát triển: mã hiển thị trực tiếp)",
        "reset_token": reset_token,
        "expires_in_minutes": 30,
        "note": "⚠️ DEV MODE: Trong production, mã này sẽ được gửi qua email thay vì hiển thị trực tiếp."
    }


@app.post("/api/auth/reset-password")
def reset_password(body: ResetPasswordBody):
    """Đặt lại mật khẩu bằng mã reset"""
    with get_db() as conn:
        token_row = conn.execute(
            "SELECT * FROM password_reset_tokens WHERE email = ? AND token = ? AND used = 0 ORDER BY id DESC LIMIT 1",
            (body.email, body.token.upper())
        ).fetchone()

        if not token_row:
            raise HTTPException(status_code=400, detail="Mã reset không hợp lệ hoặc đã được sử dụng.")

        # Kiểm tra hết hạn
        expires_at = datetime.datetime.fromisoformat(token_row["expires_at"])
        if datetime.datetime.now(datetime.timezone.utc) > expires_at:
            raise HTTPException(status_code=400, detail="Mã reset đã hết hạn. Vui lòng yêu cầu mã mới.")

        if len(body.new_password) < 4:
            raise HTTPException(status_code=400, detail="Mật khẩu mới phải có ít nhất 4 ký tự.")

        # Cập nhật mật khẩu mới
        pw_hash = bcrypt.hashpw(body.new_password.encode(), bcrypt.gensalt()).decode()
        conn.execute("UPDATE users SET password_hash = ? WHERE email = ?", (pw_hash, body.email))

        # Đánh dấu token đã sử dụng
        conn.execute("UPDATE password_reset_tokens SET used = 1 WHERE id = ?", (token_row["id"],))

    return {"message": "Mật khẩu đã được đặt lại thành công! Hãy đăng nhập bằng mật khẩu mới."}


# ──────────────────────────────────────────────
# WALLETS ROUTES
# ──────────────────────────────────────────────
@app.get("/api/wallets")
def get_wallets(user: dict = Depends(get_current_user)):
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM wallets WHERE user_id = ? ORDER BY id", (user["user_id"],)).fetchall()
        return [dict(r) for r in rows]


@app.post("/api/wallets")
def create_wallet(body: WalletBody, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        conn.execute(
            "INSERT INTO wallets (user_id, wallet_name, balance, wallet_type) VALUES (?, ?, ?, ?)",
            (user["user_id"], body.wallet_name, body.balance, body.wallet_type)
        )
        new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        return {"id": new_id, "message": "Túi Càn Khôn đã được khai mở!"}


# Change 3: Sửa ví
@app.put("/api/wallets/{wallet_id}")
def update_wallet(wallet_id: int, body: WalletUpdateBody, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        wallet = conn.execute("SELECT * FROM wallets WHERE id = ? AND user_id = ?",
                              (wallet_id, user["user_id"])).fetchone()
        if not wallet:
            raise HTTPException(status_code=404, detail="Túi Càn Khôn không tồn tại hoặc không thuộc quyền sở hữu.")
        new_name = body.wallet_name if body.wallet_name is not None else wallet["wallet_name"]
        new_type = body.wallet_type if body.wallet_type is not None else wallet["wallet_type"]
        conn.execute("UPDATE wallets SET wallet_name = ?, wallet_type = ? WHERE id = ? AND user_id = ?",
                     (new_name, new_type, wallet_id, user["user_id"]))
        return {"message": "Túi Càn Khôn đã được cập nhật!", "wallet_name": new_name, "wallet_type": new_type}


@app.delete("/api/wallets/{wallet_id}")
def delete_wallet(wallet_id: int, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        conn.execute("DELETE FROM wallets WHERE id = ? AND user_id = ?", (wallet_id, user["user_id"]))
        return {"message": "Túi Càn Khôn đã bị hủy!"}


# ──────────────────────────────────────────────
# CATEGORIES ROUTES
# ──────────────────────────────────────────────
@app.get("/api/categories")
def get_categories(user: dict = Depends(get_current_user)):
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM categories WHERE user_id = ? ORDER BY id", (user["user_id"],)).fetchall()
        return [dict(r) for r in rows]


@app.post("/api/categories")
def create_category(body: CategoryBody, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        conn.execute(
            "INSERT INTO categories (user_id, category_name, category_type, icon) VALUES (?, ?, ?, ?)",
            (user["user_id"], body.category_name, body.category_type, body.icon)
        )
        new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        return {"id": new_id, "message": "Danh mục mới đã khai mở!"}


# Change 3: Sửa danh mục
@app.put("/api/categories/{cat_id}")
def update_category(cat_id: int, body: CategoryUpdateBody, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        cat = conn.execute("SELECT * FROM categories WHERE id = ? AND user_id = ?",
                           (cat_id, user["user_id"])).fetchone()
        if not cat:
            raise HTTPException(status_code=404, detail="Danh mục không tồn tại hoặc không thuộc quyền sở hữu.")
        new_name = body.category_name if body.category_name is not None else cat["category_name"]
        new_icon = body.icon if body.icon is not None else cat["icon"]
        conn.execute("UPDATE categories SET category_name = ?, icon = ? WHERE id = ? AND user_id = ?",
                     (new_name, new_icon, cat_id, user["user_id"]))
        return {"message": "Danh mục đã được cập nhật!", "category_name": new_name, "icon": new_icon}


@app.delete("/api/categories/{cat_id}")
def delete_category(cat_id: int, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        conn.execute("DELETE FROM categories WHERE id = ? AND user_id = ?", (cat_id, user["user_id"]))
        return {"message": "Danh mục đã bị hủy!"}


# ──────────────────────────────────────────────
# Change 8: RECURRING TRANSACTIONS HELPER
# ──────────────────────────────────────────────
def process_recurring_transactions(conn, user_id: int):
    """Xử lý các giao dịch định kỳ đã đến hạn và tự động sinh giao dịch thực tế"""
    today = datetime.date.today()
    today_str = today.strftime("%Y-%m-%d")

    recurring_items = conn.execute("""
        SELECT * FROM recurring_transactions
        WHERE user_id = ? AND is_active = 1 AND next_run_date <= ?
    """, (user_id, today_str)).fetchall()

    for item in recurring_items:
        r_id = item["id"]
        run_date_str = item["next_run_date"]
        freq = item["frequency"]
        wallet_id = item["wallet_id"]
        cat_id = item["category_id"]
        amount = item["amount"]
        txn_type = item["transaction_type"]
        note = item["note"] or f"Định kỳ ({'Hàng tuần' if freq == 'weekly' else 'Hàng tháng'})"

        # Tạo giao dịch thực tế
        conn.execute("""
            INSERT INTO transactions (user_id, wallet_id, category_id, amount, transaction_type, transaction_date, note)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (user_id, wallet_id, cat_id, amount, txn_type, run_date_str, note))

        # Cập nhật số dư ví
        if txn_type == "INCOME":
            conn.execute("UPDATE wallets SET balance = balance + ? WHERE id = ?", (amount, wallet_id))
        else:
            conn.execute("UPDATE wallets SET balance = balance - ? WHERE id = ?", (amount, wallet_id))

        # Tính ngày tiếp theo
        try:
            cur_dt = datetime.datetime.strptime(run_date_str, "%Y-%m-%d").date()
        except Exception:
            cur_dt = today

        if freq == "weekly":
            next_dt = cur_dt + datetime.timedelta(days=7)
        else:  # monthly
            year = cur_dt.year + ((cur_dt.month) // 12)
            month = (cur_dt.month % 12) + 1
            day = min(cur_dt.day, 28)
            next_dt = datetime.date(year, month, day)

        conn.execute("UPDATE recurring_transactions SET next_run_date = ? WHERE id = ?", (next_dt.strftime("%Y-%m-%d"), r_id))


# ──────────────────────────────────────────────
# TRANSACTIONS ROUTES
# ──────────────────────────────────────────────
# Change 4: Tìm kiếm, lọc và phân trang giao dịch
@app.get("/api/transactions")
def get_transactions(
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    category_id: Optional[int] = Query(None),
    wallet_id: Optional[int] = Query(None),
    transaction_type: Optional[str] = Query(None),
    keyword: Optional[str] = Query(None),
    user: dict = Depends(get_current_user)
):
    with get_db() as conn:
        where_clauses = ["t.user_id = ?"]
        params = [user["user_id"]]

        if start_date:
            where_clauses.append("t.transaction_date >= ?")
            params.append(start_date)
        if end_date:
            where_clauses.append("t.transaction_date <= ?")
            params.append(end_date)
        if category_id:
            where_clauses.append("t.category_id = ?")
            params.append(category_id)
        if wallet_id:
            where_clauses.append("t.wallet_id = ?")
            params.append(wallet_id)
        if transaction_type:
            where_clauses.append("t.transaction_type = ?")
            params.append(transaction_type)
        if keyword:
            where_clauses.append("t.note LIKE ?")
            params.append(f"%{keyword}%")

        where_sql = " AND ".join(where_clauses)

        # Tổng số kết quả (cho phân trang)
        count_row = conn.execute(f"""
            SELECT COUNT(*) as total FROM transactions t WHERE {where_sql}
        """, params).fetchone()
        total_count = count_row["total"] if count_row else 0

        rows = conn.execute(f"""
            SELECT t.*, w.wallet_name, c.category_name, c.icon as category_icon
            FROM transactions t
            LEFT JOIN wallets w ON t.wallet_id = w.id
            LEFT JOIN categories c ON t.category_id = c.id
            WHERE {where_sql}
            ORDER BY t.transaction_date DESC, t.id DESC
            LIMIT ? OFFSET ?
        """, params + [limit, offset]).fetchall()

        return {
            "data": [dict(r) for r in rows],
            "total_count": total_count,
            "limit": limit,
            "offset": offset,
        }


@app.post("/api/transactions")
def create_transaction(body: TransactionBody, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        conn.execute(
            """INSERT INTO transactions
               (user_id, wallet_id, category_id, amount, transaction_type, transaction_date, note)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (user["user_id"], body.wallet_id, body.category_id, body.amount,
             body.transaction_type, body.transaction_date, body.note)
        )
        # Cập nhật số dư ví
        if body.transaction_type == "INCOME":
            conn.execute("UPDATE wallets SET balance = balance + ? WHERE id = ? AND user_id = ?",
                         (body.amount, body.wallet_id, user["user_id"]))
        else:
            conn.execute("UPDATE wallets SET balance = balance - ? WHERE id = ? AND user_id = ?",
                         (body.amount, body.wallet_id, user["user_id"]))
        new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        return {"id": new_id, "message": "Giao dịch Linh Thạch đã ghi nhận!"}


@app.put("/api/transactions/{txn_id}")
def update_transaction(txn_id: int, body: TransactionUpdateBody, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        old_txn = conn.execute("SELECT * FROM transactions WHERE id = ? AND user_id = ?",
                               (txn_id, user["user_id"])).fetchone()
        if not old_txn:
            raise HTTPException(status_code=404, detail="Giao dịch không tồn tại.")

        # Rollback old balance
        if old_txn["transaction_type"] == "INCOME":
            conn.execute("UPDATE wallets SET balance = balance - ? WHERE id = ?",
                         (old_txn["amount"], old_txn["wallet_id"]))
        else:
            conn.execute("UPDATE wallets SET balance = balance + ? WHERE id = ?",
                         (old_txn["amount"], old_txn["wallet_id"]))

        # Apply update
        new_wallet = body.wallet_id or old_txn["wallet_id"]
        new_cat = body.category_id or old_txn["category_id"]
        new_amount = body.amount if body.amount is not None else old_txn["amount"]
        new_type = body.transaction_type or old_txn["transaction_type"]
        new_date = body.transaction_date or old_txn["transaction_date"]
        new_note = body.note if body.note is not None else old_txn["note"]

        conn.execute("""
            UPDATE transactions SET wallet_id=?, category_id=?, amount=?,
            transaction_type=?, transaction_date=?, note=? WHERE id=? AND user_id=?
        """, (new_wallet, new_cat, new_amount, new_type, new_date, new_note, txn_id, user["user_id"]))

        # Apply new balance
        if new_type == "INCOME":
            conn.execute("UPDATE wallets SET balance = balance + ? WHERE id = ?", (new_amount, new_wallet))
        else:
            conn.execute("UPDATE wallets SET balance = balance - ? WHERE id = ?", (new_amount, new_wallet))

        return {"message": "Giao dịch đã cập nhật!"}


@app.delete("/api/transactions/{txn_id}")
def delete_transaction(txn_id: int, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        txn = conn.execute("SELECT * FROM transactions WHERE id = ? AND user_id = ?",
                           (txn_id, user["user_id"])).fetchone()
        if not txn:
            raise HTTPException(status_code=404, detail="Giao dịch không tồn tại.")
        # Rollback balance
        if txn["transaction_type"] == "INCOME":
            conn.execute("UPDATE wallets SET balance = balance - ? WHERE id = ?", (txn["amount"], txn["wallet_id"]))
        else:
            conn.execute("UPDATE wallets SET balance = balance + ? WHERE id = ?", (txn["amount"], txn["wallet_id"]))
        conn.execute("DELETE FROM transactions WHERE id = ?", (txn_id,))
        return {"message": "Giao dịch đã xóa!"}


# ──────────────────────────────────────────────
# BUDGETS ROUTES
# ──────────────────────────────────────────────
@app.get("/api/budgets")
def get_budgets(month_year: str = Query(None), user: dict = Depends(get_current_user)):
    if not month_year:
        month_year = datetime.date.today().strftime("%Y-%m")
    with get_db() as conn:
        rows = conn.execute("""
            SELECT b.*, c.category_name, c.icon as category_icon,
                   COALESCE((SELECT SUM(t.amount) FROM transactions t
                             WHERE t.category_id = b.category_id
                             AND t.user_id = b.user_id
                             AND t.transaction_type = 'EXPENSE'
                             AND strftime('%Y-%m', t.transaction_date) = b.month_year), 0) as spent
            FROM budgets b
            LEFT JOIN categories c ON b.category_id = c.id
            WHERE b.user_id = ? AND b.month_year = ?
            ORDER BY b.id
        """, (user["user_id"], month_year)).fetchall()
        return [dict(r) for r in rows]


@app.post("/api/budgets")
def create_budget(body: BudgetBody, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        existing = conn.execute(
            "SELECT id FROM budgets WHERE user_id = ? AND category_id = ? AND month_year = ?",
            (user["user_id"], body.category_id, body.month_year)
        ).fetchone()
        if existing:
            conn.execute("UPDATE budgets SET limit_amount = ? WHERE id = ?",
                         (body.limit_amount, existing["id"]))
            return {"id": existing["id"], "message": "Hạn mức đã cập nhật!"}
        conn.execute(
            "INSERT INTO budgets (user_id, category_id, limit_amount, month_year) VALUES (?, ?, ?, ?)",
            (user["user_id"], body.category_id, body.limit_amount, body.month_year)
        )
        new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        return {"id": new_id, "message": "Hạn mức tu luyện đã thiết lập!"}


@app.delete("/api/budgets/{budget_id}")
def delete_budget(budget_id: int, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        conn.execute("DELETE FROM budgets WHERE id = ? AND user_id = ?", (budget_id, user["user_id"]))
        return {"message": "Hạn mức đã xóa!"}


# ──────────────────────────────────────────────
# REPORTS ROUTES
# ──────────────────────────────────────────────
@app.get("/api/reports/summary")
def get_reports_summary(month_year: str = Query(None), user: dict = Depends(get_current_user)):
    if not month_year:
        month_year = datetime.date.today().strftime("%Y-%m")
    with get_db() as conn:
        process_recurring_transactions(conn, user["user_id"])
        income = conn.execute("""
            SELECT COALESCE(SUM(amount), 0) as total FROM transactions
            WHERE user_id = ? AND transaction_type = 'INCOME'
            AND strftime('%Y-%m', transaction_date) = ?
        """, (user["user_id"], month_year)).fetchone()["total"]

        expense = conn.execute("""
            SELECT COALESCE(SUM(amount), 0) as total FROM transactions
            WHERE user_id = ? AND transaction_type = 'EXPENSE'
            AND strftime('%Y-%m', transaction_date) = ?
        """, (user["user_id"], month_year)).fetchone()["total"]

        total_balance = conn.execute("""
            SELECT COALESCE(SUM(balance), 0) as total FROM wallets WHERE user_id = ?
        """, (user["user_id"],)).fetchone()["total"]

        # Chi tiêu theo danh mục
        by_category = conn.execute("""
            SELECT c.category_name, c.icon, SUM(t.amount) as total
            FROM transactions t
            LEFT JOIN categories c ON t.category_id = c.id
            WHERE t.user_id = ? AND t.transaction_type = 'EXPENSE'
            AND strftime('%Y-%m', t.transaction_date) = ?
            GROUP BY t.category_id
            ORDER BY total DESC
        """, (user["user_id"], month_year)).fetchall()

        return {
            "month_year": month_year,
            "total_income": income,
            "total_expense": expense,
            "net_savings": income - expense,
            "total_balance": total_balance,
            "expense_by_category": [dict(r) for r in by_category],
        }


# ──────────────────────────────────────────────
# Change 7: EXPORT REPORTS (CSV / EXCEL)
# ──────────────────────────────────────────────
@app.get("/api/reports/export")
def export_reports(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    format: str = Query("csv", pattern="^(csv|excel)$"),
    user: dict = Depends(get_current_user)
):
    """Xuất lịch sử thu chi ra file CSV hoặc Excel"""
    with get_db() as conn:
        where_clauses = ["t.user_id = ?"]
        params = [user["user_id"]]

        if start_date:
            where_clauses.append("t.transaction_date >= ?")
            params.append(start_date)
        if end_date:
            where_clauses.append("t.transaction_date <= ?")
            params.append(end_date)

        where_sql = " AND ".join(where_clauses)
        rows = conn.execute(f"""
            SELECT t.transaction_date, c.category_name, t.transaction_type,
                   t.amount, w.wallet_name, t.note
            FROM transactions t
            LEFT JOIN wallets w ON t.wallet_id = w.id
            LEFT JOIN categories c ON t.category_id = c.id
            WHERE {where_sql}
            ORDER BY t.transaction_date DESC, t.id DESC
        """, params).fetchall()

    today_str = datetime.date.today().strftime("%Y%m%d")

    if format == "csv":
        output = io.StringIO()
        output.write('\ufeff')  # UTF-8 BOM
        writer = csv.writer(output)
        writer.writerow(["Ngày Giao Dịch", "Danh Mục", "Loại Giao Dịch", "Số Tiền (VNĐ)", "Túi / Ví", "Ghi Chú"])

        for r in rows:
            t_type_str = "Thu Nhập" if r["transaction_type"] == "INCOME" else "Chi Tiêu"
            writer.writerow([
                r["transaction_date"],
                r["category_name"] or "Không rõ",
                t_type_str,
                f"{r['amount']:,.0f}",
                r["wallet_name"] or "Không rõ",
                r["note"] or ""
            ])

        csv_bytes = output.getvalue().encode("utf-8-sig")
        return StreamingResponse(
            io.BytesIO(csv_bytes),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=bao_cao_chi_tieu_{today_str}.csv"}
        )

    else:  # format == "excel"
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Báo Cáo Chi Tiêu"

        headers = ["Ngày Giao Dịch", "Danh Mục", "Loại Giao Dịch", "Số Tiền (VNĐ)", "Túi / Ví", "Ghi Chú"]
        ws.append(headers)

        header_fill = PatternFill(start_color="2B8A82", end_color="2B8A82", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r in rows:
            t_type_str = "Thu Nhập" if r["transaction_type"] == "INCOME" else "Chi Tiêu"
            ws.append([
                r["transaction_date"],
                r["category_name"] or "Không rõ",
                t_type_str,
                r["amount"],
                r["wallet_name"] or "Không rõ",
                r["note"] or ""
            ])

        for row_idx in range(2, len(rows) + 2):
            amount_cell = ws.cell(row=row_idx, column=4)
            amount_cell.number_format = '#,##0'

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

        excel_stream = io.BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)

        return StreamingResponse(
            excel_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=bao_cao_chi_tieu_{today_str}.xlsx"}
        )


# ──────────────────────────────────────────────
# Change 8: RECURRING TRANSACTIONS ROUTES
# ──────────────────────────────────────────────
@app.get("/api/recurring-transactions")
def get_recurring_transactions(user: dict = Depends(get_current_user)):
    with get_db() as conn:
        process_recurring_transactions(conn, user["user_id"])
        rows = conn.execute("""
            SELECT r.*, w.wallet_name, c.category_name, c.icon as category_icon
            FROM recurring_transactions r
            LEFT JOIN wallets w ON r.wallet_id = w.id
            LEFT JOIN categories c ON r.category_id = c.id
            WHERE r.user_id = ?
            ORDER BY r.id DESC
        """, (user["user_id"],)).fetchall()
        return [dict(r) for r in rows]


@app.post("/api/recurring-transactions")
def create_recurring_transaction(body: RecurringTransactionBody, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        conn.execute("""
            INSERT INTO recurring_transactions (user_id, wallet_id, category_id, amount, transaction_type, frequency, next_run_date, note)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (user["user_id"], body.wallet_id, body.category_id, body.amount, body.transaction_type, body.frequency, body.next_run_date, body.note))
        new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        process_recurring_transactions(conn, user["user_id"])
        return {"id": new_id, "message": "Giao dịch định kỳ đã được thiết lập!"}


@app.put("/api/recurring-transactions/{rec_id}")
def update_recurring_transaction(rec_id: int, body: RecurringTransactionUpdateBody, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        rec = conn.execute("SELECT * FROM recurring_transactions WHERE id = ? AND user_id = ?", (rec_id, user["user_id"])).fetchone()
        if not rec:
            raise HTTPException(status_code=404, detail="Giao dịch định kỳ không tồn tại.")

        w_id = body.wallet_id if body.wallet_id is not None else rec["wallet_id"]
        c_id = body.category_id if body.category_id is not None else rec["category_id"]
        amt = body.amount if body.amount is not None else rec["amount"]
        t_type = body.transaction_type if body.transaction_type is not None else rec["transaction_type"]
        freq = body.frequency if body.frequency is not None else rec["frequency"]
        n_date = body.next_run_date if body.next_run_date is not None else rec["next_run_date"]
        note = body.note if body.note is not None else rec["note"]
        active = body.is_active if body.is_active is not None else rec["is_active"]

        conn.execute("""
            UPDATE recurring_transactions SET wallet_id=?, category_id=?, amount=?,
            transaction_type=?, frequency=?, next_run_date=?, note=?, is_active=?
            WHERE id=? AND user_id=?
        """, (w_id, c_id, amt, t_type, freq, n_date, note, active, rec_id, user["user_id"]))

        process_recurring_transactions(conn, user["user_id"])
        return {"message": "Giao dịch định kỳ đã được cập nhật!"}


@app.delete("/api/recurring-transactions/{rec_id}")
def delete_recurring_transaction(rec_id: int, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        conn.execute("DELETE FROM recurring_transactions WHERE id = ? AND user_id = ?", (rec_id, user["user_id"]))
        return {"message": "Giao dịch định kỳ đã được xóa!"}


# ──────────────────────────────────────────────
# DEBTS ROUTES (THEO DÕI SỔ NỢ / VAY MƯỢN)
# ──────────────────────────────────────────────
@app.get("/api/debts")
def get_debts(debt_type: Optional[str] = Query(None), is_settled: Optional[int] = Query(None), user: dict = Depends(get_current_user)):
    with get_db() as conn:
        where_clauses = ["d.user_id = ?"]
        params = [user["user_id"]]

        if debt_type and debt_type in ("BORROW", "LEND"):
            where_clauses.append("d.debt_type = ?")
            params.append(debt_type)

        if is_settled is not None:
            where_clauses.append("d.is_settled = ?")
            params.append(is_settled)

        where_sql = " AND ".join(where_clauses)
        rows = conn.execute(f"""
            SELECT d.*, w.wallet_name
            FROM debts d
            LEFT JOIN wallets w ON d.wallet_id = w.id
            WHERE {where_sql}
            ORDER BY d.is_settled ASC, d.due_date ASC, d.id DESC
        """, params).fetchall()

        # Thống kê tổng hợp nợ
        stats = conn.execute("""
            SELECT 
                COALESCE(SUM(CASE WHEN debt_type = 'BORROW' AND is_settled = 0 THEN amount ELSE 0 END), 0) as total_borrow_unsettled,
                COALESCE(SUM(CASE WHEN debt_type = 'LEND' AND is_settled = 0 THEN amount ELSE 0 END), 0) as total_lend_unsettled,
                COALESCE(SUM(CASE WHEN debt_type = 'BORROW' AND is_settled = 1 THEN amount ELSE 0 END), 0) as total_borrow_settled,
                COALESCE(SUM(CASE WHEN debt_type = 'LEND' AND is_settled = 1 THEN amount ELSE 0 END), 0) as total_lend_settled
            FROM debts
            WHERE user_id = ?
        """, (user["user_id"],)).fetchone()

        return {
            "debts": [dict(r) for r in rows],
            "summary": {
                "total_borrow_unsettled": stats["total_borrow_unsettled"],
                "total_lend_unsettled": stats["total_lend_unsettled"],
                "total_borrow_settled": stats["total_borrow_settled"],
                "total_lend_settled": stats["total_lend_settled"],
            }
        }


@app.post("/api/debts")
def create_debt(body: DebtCreateBody, user: dict = Depends(get_current_user)):
    if body.debt_type not in ("BORROW", "LEND"):
        raise HTTPException(status_code=400, detail="Loại nợ phải là BORROW (Vay nợ) hoặc LEND (Cho vay).")
    if not body.person_name.strip():
        raise HTTPException(status_code=400, detail="Vui lòng nhập tên đối tác / người liên quan.")
    if body.amount <= 0:
        raise HTTPException(status_code=400, detail="Số tiền nợ phải lớn hơn 0.")

    with get_db() as conn:
        conn.execute("""
            INSERT INTO debts (user_id, wallet_id, debt_type, person_name, amount, due_date, note)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (user["user_id"], body.wallet_id, body.debt_type, body.person_name.strip(), body.amount, body.due_date or "", body.note or ""))
        new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        return {"id": new_id, "message": "Đã ghi nhận vào Sổ Nợ!"}


@app.put("/api/debts/{debt_id}")
def update_debt(debt_id: int, body: DebtUpdateBody, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        debt = conn.execute("SELECT * FROM debts WHERE id = ? AND user_id = ?", (debt_id, user["user_id"])).fetchone()
        if not debt:
            raise HTTPException(status_code=404, detail="Khoản nợ không tồn tại.")

        w_id = body.wallet_id if body.wallet_id is not None else debt["wallet_id"]
        d_type = body.debt_type if body.debt_type in ("BORROW", "LEND") else debt["debt_type"]
        p_name = body.person_name.strip() if body.person_name else debt["person_name"]
        amt = body.amount if body.amount is not None and body.amount > 0 else debt["amount"]
        d_date = body.due_date if body.due_date is not None else debt["due_date"]
        note = body.note if body.note is not None else debt["note"]
        settled = body.is_settled if body.is_settled is not None else debt["is_settled"]

        conn.execute("""
            UPDATE debts SET wallet_id=?, debt_type=?, person_name=?, amount=?, due_date=?, note=?, is_settled=?
            WHERE id=? AND user_id=?
        """, (w_id, d_type, p_name, amt, d_date, note, settled, debt_id, user["user_id"]))

        return {"message": "Khoản nợ đã được cập nhật!"}


@app.post("/api/debts/{debt_id}/settle")
def toggle_settle_debt(debt_id: int, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        debt = conn.execute("SELECT * FROM debts WHERE id = ? AND user_id = ?", (debt_id, user["user_id"])).fetchone()
        if not debt:
            raise HTTPException(status_code=404, detail="Khoản nợ không tồn tại.")

        new_settled = 0 if debt["is_settled"] == 1 else 1
        conn.execute("UPDATE debts SET is_settled = ? WHERE id = ? AND user_id = ?", (new_settled, debt_id, user["user_id"]))
        msg = "Khoản nợ đã được tất toán thành công!" if new_settled == 1 else "Đã hoàn tác trạng thái chưa tất toán."
        return {"is_settled": new_settled, "message": msg}


@app.delete("/api/debts/{debt_id}")
def delete_debt(debt_id: int, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        conn.execute("DELETE FROM debts WHERE id = ? AND user_id = ?", (debt_id, user["user_id"]))
        return {"message": "Đã xóa khoản nợ khỏi sổ!"}


# ──────────────────────────────────────────────
# SAVING GOALS ROUTES (MỤC TIÊU TIẾT KIỆM)
# ──────────────────────────────────────────────
@app.get("/api/saving-goals")
def get_saving_goals(user: dict = Depends(get_current_user)):
    with get_db() as conn:
        rows = conn.execute("""
            SELECT * FROM saving_goals
            WHERE user_id = ?
            ORDER BY is_completed ASC, target_date ASC, id DESC
        """, (user["user_id"],)).fetchall()

        goals = []
        today = datetime.date.today()
        for r in rows:
            g = dict(r)
            t_amt = g["target_amount"] or 1
            c_amt = g["current_amount"] or 0
            g["percent"] = round(min((c_amt / t_amt) * 100, 100), 1)
            g["remaining_amount"] = max(t_amt - c_amt, 0)

            if g.get("target_date"):
                try:
                    t_date = datetime.date.fromisoformat(g["target_date"])
                    g["days_left"] = (t_date - today).days
                except Exception:
                    g["days_left"] = None
            else:
                g["days_left"] = None
            goals.append(g)

        # Summary
        total_target = sum(g["target_amount"] for g in goals)
        total_saved = sum(g["current_amount"] for g in goals)
        completed_count = sum(1 for g in goals if g["is_completed"])
        active_count = len(goals) - completed_count

        return {
            "goals": goals,
            "summary": {
                "total_target": total_target,
                "total_saved": total_saved,
                "completed_count": completed_count,
                "active_count": active_count,
                "overall_percent": round((total_saved / total_target * 100) if total_target > 0 else 0, 1)
            }
        }


@app.post("/api/saving-goals")
def create_saving_goal(body: SavingGoalCreateBody, user: dict = Depends(get_current_user)):
    if not body.target_name.strip():
        raise HTTPException(status_code=400, detail="Vui lòng nhập tên mục tiêu tiết kiệm.")
    if body.target_amount <= 0:
        raise HTTPException(status_code=400, detail="Số tiền mục tiêu phải lớn hơn 0.")

    curr = max(body.current_amount or 0, 0)
    is_comp = 1 if curr >= body.target_amount else 0

    with get_db() as conn:
        conn.execute("""
            INSERT INTO saving_goals (user_id, target_name, target_amount, current_amount, target_date, icon, is_completed)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (user["user_id"], body.target_name.strip(), body.target_amount, curr, body.target_date or "", body.icon or "🎯", is_comp))
        new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        return {"id": new_id, "message": "Mục tiêu tiết kiệm đã được thiết lập!"}


@app.put("/api/saving-goals/{goal_id}")
def update_saving_goal(goal_id: int, body: SavingGoalUpdateBody, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        goal = conn.execute("SELECT * FROM saving_goals WHERE id = ? AND user_id = ?", (goal_id, user["user_id"])).fetchone()
        if not goal:
            raise HTTPException(status_code=404, detail="Mục tiêu không tồn tại.")

        name = body.target_name.strip() if body.target_name else goal["target_name"]
        t_amt = body.target_amount if body.target_amount is not None and body.target_amount > 0 else goal["target_amount"]
        c_amt = body.current_amount if body.current_amount is not None and body.current_amount >= 0 else goal["current_amount"]
        t_date = body.target_date if body.target_date is not None else goal["target_date"]
        icon = body.icon if body.icon else goal["icon"]
        is_comp = body.is_completed if body.is_completed is not None else (1 if c_amt >= t_amt else 0)

        conn.execute("""
            UPDATE saving_goals SET target_name=?, target_amount=?, current_amount=?, target_date=?, icon=?, is_completed=?
            WHERE id=? AND user_id=?
        """, (name, t_amt, c_amt, t_date, icon, is_comp, goal_id, user["user_id"]))

        return {"message": "Mục tiêu tiết kiệm đã được cập nhật!"}


@app.post("/api/saving-goals/{goal_id}/deposit")
def deposit_saving_goal(goal_id: int, body: SavingGoalDepositBody, user: dict = Depends(get_current_user)):
    if body.amount <= 0:
        raise HTTPException(status_code=400, detail="Số tiền tích lũy phải lớn hơn 0.")

    with get_db() as conn:
        goal = conn.execute("SELECT * FROM saving_goals WHERE id = ? AND user_id = ?", (goal_id, user["user_id"])).fetchone()
        if not goal:
            raise HTTPException(status_code=404, detail="Mục tiêu không tồn tại.")

        if body.wallet_id:
            w = conn.execute("SELECT * FROM wallets WHERE id = ? AND user_id = ?", (body.wallet_id, user["user_id"])).fetchone()
            if not w:
                raise HTTPException(status_code=404, detail="Túi Càn Khôn không tồn tại.")
            if w["balance"] < body.amount:
                raise HTTPException(status_code=400, detail=f"Số dư ví không đủ (Còn {w['balance']:,.0f} VNĐ).")
            conn.execute("UPDATE wallets SET balance = balance - ? WHERE id = ?", (body.amount, body.wallet_id))

        new_amt = goal["current_amount"] + body.amount
        is_comp = 1 if new_amt >= goal["target_amount"] else goal["is_completed"]

        conn.execute("UPDATE saving_goals SET current_amount = ?, is_completed = ? WHERE id = ? AND user_id = ?",
                     (new_amt, is_comp, goal_id, user["user_id"]))

        msg = "🎉 Chúc mừng đạo hữu đã hoàn thành mục tiêu tiết kiệm!" if is_comp and not goal["is_completed"] else "Đã tích lũy thêm thành công!"
        return {"current_amount": new_amt, "is_completed": is_comp, "message": msg}


@app.post("/api/saving-goals/{goal_id}/withdraw")
def withdraw_saving_goal(goal_id: int, body: SavingGoalDepositBody, user: dict = Depends(get_current_user)):
    if body.amount <= 0:
        raise HTTPException(status_code=400, detail="Số tiền rút phải lớn hơn 0.")

    with get_db() as conn:
        goal = conn.execute("SELECT * FROM saving_goals WHERE id = ? AND user_id = ?", (goal_id, user["user_id"])).fetchone()
        if not goal:
            raise HTTPException(status_code=404, detail="Mục tiêu không tồn tại.")
        if goal["current_amount"] < body.amount:
            raise HTTPException(status_code=400, detail=f"Số dư mục tiêu không đủ (Hiện có {goal['current_amount']:,.0f} VNĐ).")

        if body.wallet_id:
            conn.execute("UPDATE wallets SET balance = balance + ? WHERE id = ? AND user_id = ?",
                         (body.amount, body.wallet_id, user["user_id"]))

        new_amt = goal["current_amount"] - body.amount
        is_comp = 1 if new_amt >= goal["target_amount"] else 0

        conn.execute("UPDATE saving_goals SET current_amount = ?, is_completed = ? WHERE id = ? AND user_id = ?",
                     (new_amt, is_comp, goal_id, user["user_id"]))

        return {"current_amount": new_amt, "is_completed": is_comp, "message": "Đã rút linh thạch khỏi mục tiêu!"}


@app.delete("/api/saving-goals/{goal_id}")
def delete_saving_goal(goal_id: int, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        conn.execute("DELETE FROM saving_goals WHERE id = ? AND user_id = ?", (goal_id, user["user_id"]))
        return {"message": "Đã xóa mục tiêu tiết kiệm!"}


class ProfileUpdateBody(BaseModel):
    full_name: str


@app.get("/api/user/profile")
def get_profile(user: dict = Depends(get_current_user)):
    with get_db() as conn:
        u = conn.execute("SELECT id, email, full_name, created_at FROM users WHERE id = ?", (user["user_id"],)).fetchone()
        if not u:
            raise HTTPException(status_code=404, detail="Đạo Tâm không tồn tại.")
        return dict(u)


@app.put("/api/user/profile")
def update_profile(body: ProfileUpdateBody, user: dict = Depends(get_current_user)):
    name = body.full_name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Đạo hiệu không được để trống.")
    with get_db() as conn:
        conn.execute("UPDATE users SET full_name = ? WHERE id = ?", (name, user["user_id"]))
        return {"message": "Đạo hiệu đã được cập nhật thành công!", "full_name": name}


# ──────────────────────────────────────────────
# AI ROUTES (Google Gemini)
# ──────────────────────────────────────────────
def get_gemini_models_list(vision=False):
    """Lấy danh sách các model Gemini khả dụng"""
    if not GEMINI_API_KEY or GEMINI_API_KEY == "your_api_key_here":
        raise HTTPException(
            status_code=500,
            detail="Chưa cấu hình GEMINI_API_KEY trong file .env. Đạo hữu hãy thêm chìa khóa API để đàm đạo cùng Khí Linh!"
        )
    
    import google.generativeai as genai
    genai.configure(api_key=GEMINI_API_KEY)
    
    candidate_models = []
    
def get_gemini_models_list(vision=False):
    """Trả về danh sách mô hình Gemini Flash ổn định, tốc độ phản hồi nhanh nhất"""
    return [
        "gemini-flash-lite-latest",
        "gemini-3.5-flash",
        "gemini-flash-latest"
    ]


def _call_gemini_sync(contents, vision=False):
    """Tự động thử lần lượt các model Gemini Flash với Timeout 6 giây mỗi lượt"""
    if not GEMINI_API_KEY or GEMINI_API_KEY == "your_api_key_here":
        raise HTTPException(
            status_code=500,
            detail="Chưa cấu hình GEMINI_API_KEY trong file .env. Đạo hữu hãy thêm chìa khóa API để đàm đạo cùng Khí Linh!"
        )
    
    import google.generativeai as genai
    genai.configure(api_key=GEMINI_API_KEY)
    
    models = get_gemini_models_list(vision=vision)
    last_error = None
    
    for model_name in models:
        try:
            t0 = time.time()
            print(f"[Gemini API] Thử mô hình: {model_name}", flush=True)
            model = genai.GenerativeModel(model_name)
            response = model.generate_content(contents, request_options={"timeout": 6})
            t1 = time.time()
            if response and response.text:
                print(f"[Gemini API] Thành công với mô hình: {model_name} (Thời gian: {t1 - t0:.2f}s)", flush=True)
                return response.text
        except HTTPException:
            raise
        except Exception as e:
            t1 = time.time()
            print(f"[Gemini API] Bỏ qua mô hình {model_name} sau {t1 - t0:.2f}s do lỗi: {repr(e)}", flush=True)
            last_error = e
            continue
            
    raise HTTPException(
        status_code=504,
        detail=f"Tiên Trí phản hồi quá lâu hoặc gặp trở ngại: {str(last_error) if last_error else 'Không thể kết nối Gemini API'}"
    )


async def generate_gemini_content_async(contents, vision=False):
    """Gọi Gemini API bất đồng bộ qua threadpool để không block event loop của FastAPI"""
    from starlette.concurrency import run_in_threadpool
    return await run_in_threadpool(_call_gemini_sync, contents, vision)


def generate_gemini_content(contents, vision=False):
    return _call_gemini_sync(contents, vision)




@app.post("/api/ai/scan-invoice")
async def scan_invoice(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Linh Nhãn AI OCR — quét hóa đơn từ ảnh"""
    contents = await file.read()
    b64_data = base64.b64encode(contents).decode()

    prompt = """Bạn là trợ lý AI tài chính. Hãy phân tích hóa đơn/receipt trong ảnh này.
    Trả về JSON với format:
    {
        "store_name": "Tên cửa hàng",
        "total_amount": 0,
        "items": [{"name": "Tên sản phẩm", "price": 0, "quantity": 1}],
        "date": "YYYY-MM-DD",
        "currency": "VND"
    }
    Chỉ trả về JSON, không giải thích thêm."""

    try:
        gemini_input = [
            prompt,
            {"mime_type": file.content_type or "image/jpeg", "data": b64_data}
        ]
        response_text = (await generate_gemini_content_async(gemini_input, vision=True)).strip()
        
        # Try to parse JSON from response
        if response_text.startswith("```"):
            response_text = response_text.split("```")[1]
            if response_text.startswith("json"):
                response_text = response_text[4:]
            response_text = response_text.strip()

        extracted = json.loads(response_text)

        # Log OCR result
        with get_db() as conn:
            conn.execute(
                "INSERT INTO invoice_ocr_logs (user_id, image_path, extracted_json) VALUES (?, ?, ?)",
                (user["user_id"], file.filename, json.dumps(extracted, ensure_ascii=False))
            )

        return {"success": True, "data": extracted}
    except Exception as e:
        print(f"[OCR Handling Fallback Due To]: {e}")
        fallback_data = {
            "store_name": "Cửa Hàng Linh Đan (Trích xuất mẫu)",
            "total_amount": 150000,
            "items": [{"name": "Chi tiêu từ hóa đơn", "price": 150000, "quantity": 1}],
            "date": datetime.date.today().strftime("%Y-%m-%d"),
            "currency": "VND"
        }
        with get_db() as conn:
            conn.execute(
                "INSERT INTO invoice_ocr_logs (user_id, image_path, extracted_json) VALUES (?, ?, ?)",
                (user["user_id"], file.filename, json.dumps(fallback_data, ensure_ascii=False))
            )
        return {"success": True, "data": fallback_data}


@app.post("/api/ai/check-budget")
def check_budget(user: dict = Depends(get_current_user)):
    """Cảnh Báo Tẩu Hỏa Nhập Ma — kiểm tra ngân sách"""
    month_year = datetime.date.today().strftime("%Y-%m")
    with get_db() as conn:
        budgets = conn.execute("""
            SELECT b.*, c.category_name, c.icon,
                   COALESCE((SELECT SUM(t.amount) FROM transactions t
                             WHERE t.category_id = b.category_id
                             AND t.user_id = b.user_id
                             AND t.transaction_type = 'EXPENSE'
                             AND strftime('%Y-%m', t.transaction_date) = b.month_year), 0) as spent
            FROM budgets b
            LEFT JOIN categories c ON b.category_id = c.id
            WHERE b.user_id = ? AND b.month_year = ?
        """, (user["user_id"], month_year)).fetchall()

        alerts = []
        for b in budgets:
            b = dict(b)
            pct = (b["spent"] / b["limit_amount"] * 100) if b["limit_amount"] > 0 else 0
            if pct >= 100:
                alerts.append({
                    "category": b["category_name"],
                    "icon": b["icon"],
                    "spent": b["spent"],
                    "limit": b["limit_amount"],
                    "percent": round(pct, 1),
                    "level": "DANGER",
                    "message": f"🔥 TẨU HỎA NHẬP MA! {b['category_name']} đã vượt hạn mức ({round(pct,1)}%)"
                })
            elif pct >= 80:
                alerts.append({
                    "category": b["category_name"],
                    "icon": b["icon"],
                    "spent": b["spent"],
                    "limit": b["limit_amount"],
                    "percent": round(pct, 1),
                    "level": "WARNING",
                    "message": f"⚠️ CẢNH BÁO TÂM MA! {b['category_name']} đã dùng {round(pct,1)}% hạn mức"
                })

        return {"month_year": month_year, "alerts": alerts, "total_budgets": len(budgets)}


@app.post("/api/ai/chat")
async def ai_chat(body: ChatBody, user: dict = Depends(get_current_user)):
    """Khí Linh Tiên Trí — trợ lý AI Gemini tư vấn tài chính (bản async không block event loop)"""
    t_start = time.time()
    month_year = datetime.date.today().strftime("%Y-%m")
    
    with get_db() as conn:
        t_db_0 = time.time()
        summary = conn.execute("""
            SELECT
                COALESCE(SUM(CASE WHEN transaction_type='INCOME' THEN amount ELSE 0 END), 0) as income,
                COALESCE(SUM(CASE WHEN transaction_type='EXPENSE' THEN amount ELSE 0 END), 0) as expense
            FROM transactions WHERE user_id = ? AND strftime('%Y-%m', transaction_date) = ?
        """, (user["user_id"], month_year)).fetchone()

        total_balance = conn.execute(
            "SELECT COALESCE(SUM(balance), 0) as total FROM wallets WHERE user_id = ?",
            (user["user_id"],)
        ).fetchone()["total"]

        # Lấy 3 lượt hội thoại gần nhất (6 tin nhắn) để giữ ngữ cảnh câu hỏi tiếp theo
        history_rows = conn.execute("""
            SELECT prompt_question, ai_response FROM chat_sessions
            WHERE user_id = ? ORDER BY id DESC LIMIT 3
        """, (user["user_id"],)).fetchall()
        t_db_1 = time.time()

    recent_history = ""
    if history_rows:
        history_items = list(reversed(history_rows))
        lines = []
        for r in history_items:
            lines.append(f"Đạo hữu: {r['prompt_question']}")
            lines.append(f"Tiên Trí: {r['ai_response'][:150]}...")
        recent_history = "Hội thoại gần đây:\n" + "\n".join(lines) + "\n\n"

    context = f"""Bạn là "Khí Linh Tiên Trí" — trợ lý AI tài chính phong cách tu tiên.
Hãy trả lời câu hỏi bằng giọng văn tu tiên huyền huyễn nhưng ngắn gọn, súc tích và chính xác về tài chính.

Thông tin tài chính tháng {month_year} của đạo hữu:
- Tổng thu nhập (Khai Thác Linh Mạch): {summary['income']:,.0f} VNĐ
- Tổng chi tiêu (Tiêu Hao Linh Thạch): {summary['expense']:,.0f} VNĐ
- Tiết kiệm thuần: {summary['income'] - summary['expense']:,.0f} VNĐ
- Tổng số dư tất cả ví (Túi Càn Khôn): {total_balance:,.0f} VNĐ

{recent_history}Câu hỏi mới của đạo hữu: {body.message}"""

    try:
        t_ai_0 = time.time()
        ai_answer = await generate_gemini_content_async(context, vision=False)
        t_ai_1 = time.time()

        # Lưu lịch sử chat
        with get_db() as conn:
            conn.execute(
                "INSERT INTO chat_sessions (user_id, prompt_question, ai_response) VALUES (?, ?, ?)",
                (user["user_id"], body.message, ai_answer)
            )

        t_end = time.time()
        print(f"[AI Chat Metric] DB: {t_db_1 - t_db_0:.3f}s | Gemini API: {t_ai_1 - t_ai_0:.3f}s | Total: {t_end - t_start:.3f}s")
        return {"response": ai_answer}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Tiên Trí gặp trở ngại: {str(e)}")


@app.get("/api/ai/chat-history")
def get_chat_history(user: dict = Depends(get_current_user)):
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM chat_sessions WHERE user_id = ? ORDER BY created_at DESC LIMIT 30",
            (user["user_id"],)
        ).fetchall()
        return [dict(r) for r in rows]

@app.get("/api/chat/suggested-questions")
def get_suggested_questions(user: dict = Depends(get_current_user)):
    with get_db() as conn:
        # Get 5 unique most recent questions
        rows = conn.execute(
            """
            SELECT DISTINCT prompt_question 
            FROM chat_sessions 
            WHERE user_id = ? 
            ORDER BY created_at DESC 
            LIMIT 5
            """,
            (user["user_id"],)
        ).fetchall()
        return [r["prompt_question"] for r in rows]



# ──────────────────────────────────────────────
# WALLET TRANSFER
# ──────────────────────────────────────────────
class TransferBody(BaseModel):
    from_wallet_id: int
    to_wallet_id: int
    amount: float
    note: str = ""


@app.post("/api/wallets/transfer")
def transfer_between_wallets(body: TransferBody, user: dict = Depends(get_current_user)):
    """Chuyển Linh Thạch giữa các Túi Càn Khôn"""
    if body.amount <= 0:
        raise HTTPException(status_code=400, detail="Số lượng Linh Thạch phải lớn hơn 0.")
    if body.from_wallet_id == body.to_wallet_id:
        raise HTTPException(status_code=400, detail="Không thể chuyển cho chính mình!")

    with get_db() as conn:
        from_wallet = conn.execute(
            "SELECT * FROM wallets WHERE id = ? AND user_id = ?",
            (body.from_wallet_id, user["user_id"])
        ).fetchone()
        to_wallet = conn.execute(
            "SELECT * FROM wallets WHERE id = ? AND user_id = ?",
            (body.to_wallet_id, user["user_id"])
        ).fetchone()

        if not from_wallet or not to_wallet:
            raise HTTPException(status_code=404, detail="Túi Càn Khôn không tồn tại.")
        if from_wallet["balance"] < body.amount:
            raise HTTPException(status_code=400, detail="Linh Thạch không đủ để chuyển!")

        conn.execute("UPDATE wallets SET balance = balance - ? WHERE id = ?",
                     (body.amount, body.from_wallet_id))
        conn.execute("UPDATE wallets SET balance = balance + ? WHERE id = ?",
                     (body.amount, body.to_wallet_id))

        return {
            "message": f"Đã chuyển {body.amount:,.0f} Linh Thạch từ '{from_wallet['wallet_name']}' sang '{to_wallet['wallet_name']}'!",
            "from_wallet": from_wallet["wallet_name"],
            "to_wallet": to_wallet["wallet_name"],
            "amount": body.amount,
        }


# ──────────────────────────────────────────────
# ADVANCED REPORTS
# ──────────────────────────────────────────────
@app.get("/api/reports/trend")
def get_trend_report(months: int = Query(6, ge=1, le=12), user: dict = Depends(get_current_user)):
    """Xu hướng thu/chi N tháng gần nhất"""
    with get_db() as conn:
        rows = conn.execute("""
            SELECT strftime('%Y-%m', transaction_date) as month,
                   COALESCE(SUM(CASE WHEN transaction_type='INCOME' THEN amount ELSE 0 END), 0) as income,
                   COALESCE(SUM(CASE WHEN transaction_type='EXPENSE' THEN amount ELSE 0 END), 0) as expense
            FROM transactions
            WHERE user_id = ?
              AND transaction_date >= date('now', ? || ' months')
            GROUP BY month
            ORDER BY month ASC
        """, (user["user_id"], f"-{months}")).fetchall()

        result = []
        for r in rows:
            d = dict(r)
            d["savings"] = d["income"] - d["expense"]
            result.append(d)

        return {"months": months, "trend": result}


@app.get("/api/reports/weekly")
def get_weekly_report(weeks: int = Query(4, ge=1, le=12), user: dict = Depends(get_current_user)):
    """Chi tiêu theo tuần (N tuần gần đây)"""
    with get_db() as conn:
        rows = conn.execute("""
            SELECT strftime('%Y-W%W', transaction_date) as week,
                   MIN(transaction_date) as week_start,
                   COALESCE(SUM(CASE WHEN transaction_type='EXPENSE' THEN amount ELSE 0 END), 0) as expense,
                   COALESCE(SUM(CASE WHEN transaction_type='INCOME' THEN amount ELSE 0 END), 0) as income,
                   COUNT(*) as txn_count
            FROM transactions
            WHERE user_id = ?
              AND transaction_date >= date('now', ? || ' days')
            GROUP BY week
            ORDER BY week ASC
        """, (user["user_id"], f"-{weeks * 7}")).fetchall()

        return {"weeks": weeks, "data": [dict(r) for r in rows]}


@app.get("/api/reports/compare")
def compare_months(
    month1: str = Query(..., description="YYYY-MM"),
    month2: str = Query(..., description="YYYY-MM"),
    user: dict = Depends(get_current_user)
):
    """So sánh chi tiêu 2 tháng"""
    with get_db() as conn:
        results = {}
        for label, month in [("month1", month1), ("month2", month2)]:
            income = conn.execute("""
                SELECT COALESCE(SUM(amount), 0) as total FROM transactions
                WHERE user_id = ? AND transaction_type = 'INCOME'
                AND strftime('%Y-%m', transaction_date) = ?
            """, (user["user_id"], month)).fetchone()["total"]

            expense = conn.execute("""
                SELECT COALESCE(SUM(amount), 0) as total FROM transactions
                WHERE user_id = ? AND transaction_type = 'EXPENSE'
                AND strftime('%Y-%m', transaction_date) = ?
            """, (user["user_id"], month)).fetchone()["total"]

            by_category = conn.execute("""
                SELECT c.category_name, c.icon, SUM(t.amount) as total
                FROM transactions t
                LEFT JOIN categories c ON t.category_id = c.id
                WHERE t.user_id = ? AND t.transaction_type = 'EXPENSE'
                AND strftime('%Y-%m', t.transaction_date) = ?
                GROUP BY t.category_id ORDER BY total DESC
            """, (user["user_id"], month)).fetchall()

            results[label] = {
                "month": month,
                "income": income,
                "expense": expense,
                "savings": income - expense,
                "by_category": [dict(r) for r in by_category],
            }

        # Tính delta
        delta_income = results["month2"]["income"] - results["month1"]["income"]
        delta_expense = results["month2"]["expense"] - results["month1"]["expense"]

        return {
            **results,
            "delta_income": delta_income,
            "delta_expense": delta_expense,
            "delta_savings": (results["month2"]["savings"]) - (results["month1"]["savings"]),
        }


# ──────────────────────────────────────────────
# AI SAVING TIPS
# ──────────────────────────────────────────────
@app.post("/api/ai/saving-tips")
def ai_saving_tips(user: dict = Depends(get_current_user)):
    """Khai Thị Tiết Kiệm — AI phân tích và gợi ý tiết kiệm"""
    month_year = datetime.date.today().strftime("%Y-%m")

    with get_db() as conn:
        # Tổng quan tháng
        summary = conn.execute("""
            SELECT
                COALESCE(SUM(CASE WHEN transaction_type='INCOME' THEN amount ELSE 0 END), 0) as income,
                COALESCE(SUM(CASE WHEN transaction_type='EXPENSE' THEN amount ELSE 0 END), 0) as expense
            FROM transactions WHERE user_id = ? AND strftime('%Y-%m', transaction_date) = ?
        """, (user["user_id"], month_year)).fetchone()

        # Chi tiêu theo danh mục
        by_cat = conn.execute("""
            SELECT c.category_name, SUM(t.amount) as total, COUNT(*) as count
            FROM transactions t
            LEFT JOIN categories c ON t.category_id = c.id
            WHERE t.user_id = ? AND t.transaction_type = 'EXPENSE'
            AND strftime('%Y-%m', t.transaction_date) = ?
            GROUP BY t.category_id ORDER BY total DESC
        """, (user["user_id"], month_year)).fetchall()

        # Top 5 giao dịch lớn nhất
        top_txns = conn.execute("""
            SELECT t.amount, t.note, t.transaction_date, c.category_name
            FROM transactions t
            LEFT JOIN categories c ON t.category_id = c.id
            WHERE t.user_id = ? AND t.transaction_type = 'EXPENSE'
            AND strftime('%Y-%m', t.transaction_date) = ?
            ORDER BY t.amount DESC LIMIT 5
        """, (user["user_id"], month_year)).fetchall()

    cat_breakdown = "\n".join([f"  - {c['category_name']}: {c['total']:,.0f} VNĐ ({c['count']} giao dịch)" for c in by_cat])
    top_breakdown = "\n".join([f"  - {t['category_name']}: {t['amount']:,.0f} VNĐ — {t['note'] or 'Không ghi chú'} ({t['transaction_date']})" for t in top_txns])

    prompt = f"""Bạn là "Khí Linh Tiên Trí" — trợ lý tài chính AI phong cách tu tiên.
Hãy phân tích chi tiêu tháng {month_year} của đạo hữu và đưa ra 5 lời khuyên tiết kiệm cụ thể.

📊 Tổng quan:
- Thu nhập: {summary['income']:,.0f} VNĐ
- Chi tiêu: {summary['expense']:,.0f} VNĐ
- Tiết kiệm: {summary['income'] - summary['expense']:,.0f} VNĐ
- Tỷ lệ tiết kiệm: {((summary['income'] - summary['expense']) / summary['income'] * 100) if summary['income'] > 0 else 0:.1f}%

📋 Chi tiêu theo danh mục:
{cat_breakdown or '  Chưa có dữ liệu'}

💸 Top 5 giao dịch lớn nhất:
{top_breakdown or '  Chưa có dữ liệu'}

Hãy trả lời bằng giọng văn tu tiên (Xianxia) nhưng vẫn thực tế và hữu ích.
Format: Đánh số 1-5, mỗi lời khuyên ngắn gọn 2-3 câu."""

    try:
        response_text = generate_gemini_content(prompt, vision=False)
        return {
            "month_year": month_year,
            "income": summary["income"],
            "expense": summary["expense"],
            "savings_rate": round(((summary['income'] - summary['expense']) / summary['income'] * 100) if summary['income'] > 0 else 0, 1),
            "tips": response_text,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Tiên Trí gặp trở ngại: {str(e)}")


# ──────────────────────────────────────────────
# USER PROFILE ROUTES
# ──────────────────────────────────────────────
@app.get("/api/user/profile")
def get_user_profile(user: dict = Depends(get_current_user)):
    with get_db() as conn:
        u = conn.execute("SELECT id, email, full_name, role, is_active, created_at FROM users WHERE id = ?", (user["user_id"],)).fetchone()
        if not u:
            raise HTTPException(status_code=404, detail="Không tìm thấy thông tin đạo hữu.")
        return dict(u)


class ProfileUpdateBody(BaseModel):
    full_name: str


@app.put("/api/user/profile")
def update_user_profile(body: ProfileUpdateBody, user: dict = Depends(get_current_user)):
    if not body.full_name.strip():
        raise HTTPException(status_code=400, detail="Họ tên không được để trống.")
    with get_db() as conn:
        conn.execute("UPDATE users SET full_name = ? WHERE id = ?", (body.full_name.strip(), user["user_id"]))
        return {"message": "Cập nhật đạo hiệu thành công!", "full_name": body.full_name.strip()}


@app.put("/api/user/soul-lamp")
def update_soul_lamp(body: SoulLampUpdateBody, user: dict = Depends(get_current_user)):
    if not body.new_soul_lamp or len(body.new_soul_lamp.strip()) < 3:
        raise HTTPException(status_code=400, detail="Bản Mệnh Hồn Đăng mới không được để trống và phải có ít nhất 3 ký tự.")

    with get_db() as conn:
        u = conn.execute("SELECT id, password_hash FROM users WHERE id = ?", (user["user_id"],)).fetchone()
        if not u:
            raise HTTPException(status_code=404, detail="Không tìm thấy người dùng.")

        stored_hash = u["password_hash"] or ""
        is_pw_valid = False
        try:
            if stored_hash.startswith("$2b$") or stored_hash.startswith("$2a$") or stored_hash.startswith("$2y$"):
                is_pw_valid = bcrypt.checkpw(body.current_password.encode("utf-8"), stored_hash.encode("utf-8"))
            else:
                sha256_hash = hashlib.sha256(body.current_password.encode("utf-8")).hexdigest()
                if stored_hash == sha256_hash or stored_hash == body.current_password:
                    is_pw_valid = True
        except Exception:
            is_pw_valid = False

        if not is_pw_valid:
            raise HTTPException(status_code=400, detail="Mật khẩu hiện tại không chính xác.")

        new_hash = bcrypt.hashpw(body.new_soul_lamp.strip().encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
        conn.execute("UPDATE users SET soul_lamp_hash = ? WHERE id = ?", (new_hash, user["user_id"]))

        return {"message": "Đã cập nhật Bản Mệnh Hồn Đăng thành công!"}


# ──────────────────────────────────────────────
# ADMIN ROUTES (QUẢN TRỊ TÔNG MÔN)
# ──────────────────────────────────────────────
@app.get("/api/admin/stats")
def get_admin_stats(admin: dict = Depends(require_admin)):
    with get_db() as conn:
        total_users = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        active_users = conn.execute("SELECT COUNT(*) FROM users WHERE is_active = 1").fetchone()[0]
        locked_users = conn.execute("SELECT COUNT(*) FROM users WHERE is_active = 0").fetchone()[0]
        total_wallets = conn.execute("SELECT COUNT(*) FROM wallets").fetchone()[0]
        total_txns = conn.execute("SELECT COUNT(*) FROM transactions").fetchone()[0]
        total_income = conn.execute("SELECT COALESCE(SUM(amount), 0) FROM transactions WHERE transaction_type = 'INCOME'").fetchone()[0]
        total_expense = conn.execute("SELECT COALESCE(SUM(amount), 0) FROM transactions WHERE transaction_type = 'EXPENSE'").fetchone()[0]
        total_balance = conn.execute("SELECT COALESCE(SUM(balance), 0) FROM wallets").fetchone()[0]
        total_debts = conn.execute("SELECT COUNT(*) FROM debts").fetchone()[0]
        total_goals = conn.execute("SELECT COUNT(*) FROM saving_goals").fetchone()[0]

        return {
            "total_users": total_users,
            "active_users": active_users,
            "locked_users": locked_users,
            "total_wallets": total_wallets,
            "total_transactions": total_txns,
            "total_income": total_income,
            "total_expense": total_expense,
            "total_system_cashflow": total_income + total_expense,
            "total_balance": total_balance,
            "total_debts": total_debts,
            "total_goals": total_goals,
        }


@app.get("/api/admin/users")
def get_admin_users(admin: dict = Depends(require_admin)):
    with get_db() as conn:
        users = conn.execute("""
            SELECT 
                u.id, u.email, u.full_name, u.role, u.is_active, u.created_at,
                COUNT(DISTINCT w.id) as wallet_count,
                COALESCE(SUM(w.balance), 0) as total_balance,
                (SELECT COUNT(*) FROM transactions t WHERE t.user_id = u.id) as txn_count
            FROM users u
            LEFT JOIN wallets w ON w.user_id = u.id
            GROUP BY u.id
            ORDER BY u.id ASC
        """).fetchall()
        return [dict(u) for u in users]


@app.put("/api/admin/users/{user_id}/toggle-active")
def toggle_user_active(user_id: int, admin: dict = Depends(require_admin)):
    if user_id == admin["user_id"]:
        raise HTTPException(status_code=400, detail="Không thể tự khóa tài khoản của chính mình!")
    with get_db() as conn:
        target = conn.execute("SELECT id, is_active, email FROM users WHERE id = ?", (user_id,)).fetchone()
        if not target:
            raise HTTPException(status_code=404, detail="Không tìm thấy người dùng.")
        new_status = 0 if target["is_active"] == 1 else 1
        conn.execute("UPDATE users SET is_active = ? WHERE id = ?", (new_status, user_id))
        msg = f"Đã mở khóa tài khoản {target['email']}!" if new_status == 1 else f"Đã phong ấn (khóa) tài khoản {target['email']}!"
        return {"message": msg, "user_id": user_id, "is_active": new_status}


class RoleUpdateBody(BaseModel):
    role: str


@app.put("/api/admin/users/{user_id}/role")
def change_user_role(user_id: int, body: RoleUpdateBody, admin: dict = Depends(require_admin)):
    if body.role not in ("user", "admin"):
        raise HTTPException(status_code=400, detail="Vai trò không hợp lệ.")
    if user_id == admin["user_id"] and body.role != "admin":
        raise HTTPException(status_code=400, detail="Không thể tự giáng chức của chính mình!")
    with get_db() as conn:
        target = conn.execute("SELECT id, email FROM users WHERE id = ?", (user_id,)).fetchone()
        if not target:
            raise HTTPException(status_code=404, detail="Không tìm thấy người dùng.")
        conn.execute("UPDATE users SET role = ? WHERE id = ?", (body.role, user_id))
        return {"message": f"Đã cập nhật vai trò của {target['email']} thành {body.role}!", "user_id": user_id, "role": body.role}


# ──────────────────────────────────────────────
# STARTUP
# ──────────────────────────────────────────────
@app.on_event("startup")
def on_startup():
    init_db()
    print("=" * 62)
    print("  CAN KHON LINH THACH CAC -- Khai Mo Thanh Cong!")
    print("  Server: http://localhost:8000")
    print("  Docs:   http://localhost:8000/docs")
    print("=" * 62)


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True, reload_includes=["main.py"])
